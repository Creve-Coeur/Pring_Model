from __future__ import annotations

import html
import io
import re
from datetime import date, datetime, time
from pathlib import Path
from typing import Iterable

from docx import Document
from docx.document import Document as DocxDocument
from docx.oxml.ns import qn
from docx.table import Table, _Cell
from docx.text.paragraph import Paragraph
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent
ASSET_DIR = ROOT / "report_assets"
OUTPUT_FILE = ROOT / "index.html"

DOCS = [
    {
        "id": "main",
        "label": "主报告",
        "filename": "普林格新四核轮动模型投资报告.docx",
    },
    {
        "id": "theory",
        "label": "理论补充",
        "filename": "普林格理论相关.docx",
    },
    {
        "id": "factors",
        "label": "因子依据",
        "filename": "普林格周期判断的因子依据.docx",
    },
    {
        "id": "convertible",
        "label": "可转债研究",
        "filename": "可转债资产研究报告.docx",
    },
]

FILENAME_TO_ID = {item["filename"]: item["id"] for item in DOCS}
DOC_BY_ID = {item["id"]: item for item in DOCS}


def esc(value: object) -> str:
    return html.escape(str(value), quote=True)


def clean_id(value: str, fallback: str) -> str:
    value = re.sub(r"[^a-zA-Z0-9_-]+", "-", value.strip().lower()).strip("-")
    return value or fallback


def iter_blocks(parent: DocxDocument | _Cell) -> Iterable[Paragraph | Table]:
    if isinstance(parent, DocxDocument):
        parent_element = parent.element.body
    else:
        parent_element = parent._tc

    for child in parent_element.iterchildren():
        if child.tag == qn("w:p"):
            yield Paragraph(child, parent)
        elif child.tag == qn("w:tbl"):
            yield Table(child, parent)


def outline_level(paragraph: Paragraph) -> int | None:
    ppr = paragraph._p.pPr
    if ppr is None:
        return None
    node = ppr.find(qn("w:outlineLvl"))
    if node is None:
        return None
    value = node.get(qn("w:val"))
    return int(value) if value is not None else None


def num_level(paragraph: Paragraph) -> int | None:
    ppr = paragraph._p.pPr
    if ppr is None or ppr.numPr is None:
        return None
    ilvl = ppr.numPr.ilvl
    if ilvl is None:
        return 0
    return int(ilvl.val)


def alignment_class(paragraph: Paragraph) -> str:
    value = paragraph.alignment
    if value is None:
        return ""
    try:
        code = int(value)
    except TypeError:
        return ""
    return {
        1: " align-center",
        2: " align-right",
        3: " align-justify",
    }.get(code, "")


def save_image_assets(doc: DocxDocument, doc_id: str) -> dict[str, str]:
    image_dir = ASSET_DIR / doc_id
    image_dir.mkdir(parents=True, exist_ok=True)
    result: dict[str, str] = {}

    for rel in doc.part.rels.values():
        if "image" not in rel.reltype:
            continue
        part_name = Path(str(rel.target_part.partname)).name
        target = image_dir / part_name
        target.write_bytes(rel.target_part.blob)
        result[rel.rId] = target.relative_to(ROOT).as_posix()

    return result


def drawing_data(drawing, image_map: dict[str, str]) -> list[dict[str, object]]:
    images = []
    blips = drawing.xpath(".//a:blip")
    extents = drawing.xpath(".//wp:extent")
    doc_props = drawing.xpath(".//wp:docPr")

    for index, blip in enumerate(blips):
        rid = blip.get(qn("r:embed")) or blip.get(qn("r:link"))
        if not rid or rid not in image_map:
            continue

        width = None
        height = None
        if extents:
            extent = extents[min(index, len(extents) - 1)]
            cx = extent.get("cx")
            cy = extent.get("cy")
            if cx and cy:
                width = round(int(cx) / 914400 * 96)
                height = round(int(cy) / 914400 * 96)

        alt = "报告图片"
        if doc_props:
            prop = doc_props[min(index, len(doc_props) - 1)]
            alt = prop.get("descr") or prop.get("name") or alt

        images.append(
            {
                "src": image_map[rid],
                "width": width,
                "height": height,
                "alt": alt,
            }
        )

    return images


def paragraph_images(paragraph: Paragraph, image_map: dict[str, str]) -> list[dict[str, object]]:
    images = []
    for drawing in paragraph._p.xpath(".//w:drawing"):
        images.extend(drawing_data(drawing, image_map))
    return images


def image_html(image: dict[str, object]) -> str:
    style = ""
    if image.get("width"):
        style = f' style="--image-width: {int(image["width"])}px;"'
    width_attr = f' width="{int(image["width"])}"' if image.get("width") else ""
    height_attr = f' height="{int(image["height"])}"' if image.get("height") else ""
    src = esc(image["src"])
    alt = esc(image["alt"])
    return (
        f'<button class="image-zoom" type="button" data-image-src="{src}" aria-label="查看图片"{style}>'
        f'<img src="{src}" alt="{alt}" loading="lazy"{width_attr}{height_attr}>'
        "</button>"
    )


def run_html(run) -> str:
    text = run.text
    if not text:
        return ""

    piece = esc(text).replace("\n", "<br>")
    styles = []

    color = run.font.color.rgb if run.font.color and run.font.color.rgb else None
    if color:
        styles.append(f"color: #{color};")

    if run.italic:
        piece = f"<em>{piece}</em>"
    if run.underline:
        piece = f"<u>{piece}</u>"
    if run.bold:
        piece = f"<strong>{piece}</strong>"

    if styles:
        return f'<span style="{" ".join(styles)}">{piece}</span>'
    return piece


def paragraph_inner_html(paragraph: Paragraph) -> str:
    return "".join(run_html(run) for run in paragraph.runs)


def render_table(table: Table) -> str:
    rows = []
    for row_index, row in enumerate(table.rows):
        cells = []
        tag = "th" if row_index == 0 else "td"
        for cell in row.cells:
            text = esc(cell.text).replace("\n", "<br>")
            cells.append(f"<{tag}>{text}</{tag}>")
        rows.append("<tr>" + "".join(cells) + "</tr>")
    return '<div class="table-wrap"><table>' + "".join(rows) + "</table></div>"


def embedded_workbook_blobs(paragraph: Paragraph) -> list[dict[str, object]]:
    blobs = []
    for node in paragraph._p.xpath(".//*[local-name()='OLEObject']"):
        rid = node.get(qn("r:id"))
        if not rid or rid not in paragraph.part.rels:
            continue
        rel = paragraph.part.rels[rid]
        target_name = Path(str(rel.target_ref)).name
        if not target_name.lower().endswith(".xlsx"):
            continue
        blobs.append(
            {
                "name": target_name,
                "blob": rel.target_part.blob,
            }
        )
    return blobs


def decimal_places(number_format: str, default: int = 2) -> int:
    match = re.search(r"0\.([0#]+)%", number_format)
    if match:
        return len(match.group(1))
    if "%" in number_format:
        return 0
    return default


def format_workbook_value(value, number_format: str) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, time):
        return value.strftime("%H:%M")
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if "%" in number_format:
            digits = decimal_places(number_format)
            return f"{value * 100:.{digits}f}%"
        if isinstance(value, float):
            if value.is_integer():
                return str(int(value))
            return f"{value:.4f}".rstrip("0").rstrip(".")
        return str(value)
    return str(value)


def cell_inline_style(cell) -> str:
    styles = []

    horizontal = cell.alignment.horizontal
    if horizontal in {"center", "right", "left"}:
        styles.append(f"text-align: {horizontal};")

    if cell.font and cell.font.bold:
        styles.append("font-weight: 800;")

    fill = cell.fill
    color = None
    if fill and fill.fill_type and fill.fgColor and fill.fgColor.type == "rgb":
        rgb = fill.fgColor.rgb
        if rgb and rgb not in {"00000000", "FFFFFFFF"}:
            color = "#" + rgb[-6:]
    if color:
        styles.append(f"background: {color};")

    return f' style="{" ".join(styles)}"' if styles else ""


def workbook_used_bounds(ws) -> tuple[int, int, int, int] | None:
    rows = []
    cols = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                rows.append(cell.row)
                cols.append(cell.column)
    if not rows:
        return None
    return min(rows), min(cols), max(rows), max(cols)


def render_worksheet(ws) -> str:
    bounds = workbook_used_bounds(ws)
    if not bounds:
        return ""

    min_row, min_col, max_row, max_col = bounds
    merges = {}
    skipped = set()
    for merged_range in ws.merged_cells.ranges:
        if (
            merged_range.max_row < min_row
            or merged_range.min_row > max_row
            or merged_range.max_col < min_col
            or merged_range.min_col > max_col
        ):
            continue
        top_left = (merged_range.min_row, merged_range.min_col)
        merges[top_left] = {
            "rowspan": merged_range.max_row - merged_range.min_row + 1,
            "colspan": merged_range.max_col - merged_range.min_col + 1,
        }
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                if (row, col) != top_left:
                    skipped.add((row, col))

    col_group = []
    for col in range(min_col, max_col + 1):
        letter = get_column_letter(col)
        width = ws.column_dimensions[letter].width
        width_px = max(86, min(260, round((width or 12) * 8)))
        col_group.append(f'<col style="width: {width_px}px;">')

    rows = []
    for row_index in range(min_row, max_row + 1):
        cells = []
        tag = "th" if row_index == min_row else "td"
        for col_index in range(min_col, max_col + 1):
            if (row_index, col_index) in skipped:
                continue
            cell = ws.cell(row_index, col_index)
            merge_attrs = ""
            if (row_index, col_index) in merges:
                merge = merges[(row_index, col_index)]
                merge_attrs = f' rowspan="{merge["rowspan"]}" colspan="{merge["colspan"]}"'
            value = esc(format_workbook_value(cell.value, cell.number_format))
            cells.append(f"<{tag}{merge_attrs}{cell_inline_style(cell)}>{value}</{tag}>")
        rows.append("<tr>" + "".join(cells) + "</tr>")

    return (
        '<div class="embedded-sheet">'
        f'<div class="embedded-sheet-name">{esc(ws.title)}</div>'
        '<div class="table-wrap embedded-table-wrap">'
        '<table class="embedded-table">'
        f'<colgroup>{"".join(col_group)}</colgroup>'
        + "".join(rows)
        + "</table></div></div>"
    )


def render_embedded_workbook(workbook_blob: bytes, doc_id: str, workbook_name: str) -> str:
    workbook_dir = ASSET_DIR / doc_id / "workbooks"
    workbook_dir.mkdir(parents=True, exist_ok=True)
    workbook_path = workbook_dir / workbook_name
    workbook_path.write_bytes(workbook_blob)
    download_href = workbook_path.relative_to(ROOT).as_posix()

    workbook = load_workbook(io.BytesIO(workbook_blob), data_only=True)
    sheets_html = "".join(render_worksheet(ws) for ws in workbook.worksheets)
    if not sheets_html:
        return ""

    return f"""
<section class="embedded-workbook">
  <div class="embedded-workbook-head">
    <span>完整电子表格</span>
    <a href="{esc(download_href)}" download>下载原始 Excel</a>
  </div>
  {sheets_html}
</section>
""".strip()


def render_callout(prompt: str, target_file: str) -> str:
    target_id = FILENAME_TO_ID[target_file]
    target = DOC_BY_ID[target_id]
    button_label = Path(target_file).stem
    prompt_text = prompt.replace("%%%", "", 1).strip()
    return f"""
<aside class="jump-callout">
  <div class="jump-copy">{esc(prompt_text)}</div>
  <a class="jump-button cross-report-link" href="#{esc(target_id)}-title" data-target-report="{esc(target_id)}">
    <span>查看《{esc(button_label)}》</span>
  </a>
</aside>
""".strip()


def render_paragraph(
    paragraph: Paragraph,
    doc_id: str,
    image_map: dict[str, str],
    state: dict[str, object],
) -> tuple[str, dict[str, object] | None]:
    text = paragraph.text.strip()
    images = paragraph_images(paragraph, image_map)
    workbooks = embedded_workbook_blobs(paragraph)
    if workbooks:
        html_parts = []
        for workbook_index, workbook in enumerate(workbooks, start=1):
            workbook_name = f"{state['embedded_index']}-{workbook['name']}"
            state["embedded_index"] = int(state["embedded_index"]) + 1
            rendered = render_embedded_workbook(
                workbook["blob"],
                doc_id,
                workbook_name,
            )
            if rendered:
                html_parts.append(rendered)
        return "\n".join(html_parts), None

    if not text and not images:
        return "", None

    if not state["saw_title"]:
        state["saw_title"] = True
        heading_id = f"{doc_id}-title"
        entry = {"id": heading_id, "text": text, "level": 0}
        inner = paragraph_inner_html(paragraph) or esc(text)
        return f'<h1 id="{heading_id}" class="report-title">{inner}</h1>', entry

    outline = outline_level(paragraph)
    if outline is not None and text:
        heading_index = int(state["heading_index"])
        state["heading_index"] = heading_index + 1
        heading_id = f"{doc_id}-h-{heading_index}"
        html_level = min(outline + 2, 5)
        entry = {"id": heading_id, "text": text, "level": outline + 1}
        inner = paragraph_inner_html(paragraph) or esc(text)
        return f'<h{html_level} id="{heading_id}">{inner}</h{html_level}>', entry

    if images:
        body = "".join(image_html(image) for image in images)
        return f'<figure class="report-figure">{body}</figure>', None

    classes = ["report-paragraph"]
    align = alignment_class(paragraph).strip()
    if align:
        classes.append(align)
    level = num_level(paragraph)
    style = ""
    if level is not None:
        classes.append("list-paragraph")
        style = f' style="--list-indent: {level};"'
    inner = paragraph_inner_html(paragraph) or esc(text)
    return f'<p class="{" ".join(classes)}"{style}>{inner}</p>', None


def convert_doc(doc_info: dict[str, str]) -> dict[str, object]:
    doc_id = doc_info["id"]
    doc_path = ROOT / doc_info["filename"]
    document = Document(str(doc_path))
    image_map = save_image_assets(document, doc_id)

    blocks = list(iter_blocks(document))
    state: dict[str, object] = {"saw_title": False, "heading_index": 1, "embedded_index": 1}
    html_blocks: list[str] = []
    toc: list[dict[str, object]] = []

    index = 0
    while index < len(blocks):
        block = blocks[index]

        if isinstance(block, Paragraph):
            text = block.text.strip()
            next_text = ""
            if index + 1 < len(blocks) and isinstance(blocks[index + 1], Paragraph):
                next_text = blocks[index + 1].text.strip()

            if doc_id == "main" and text.startswith("%%%") and next_text in FILENAME_TO_ID:
                html_blocks.append(render_callout(text, next_text))
                index += 2
                continue

            block_html, entry = render_paragraph(block, doc_id, image_map, state)
            if block_html:
                html_blocks.append(block_html)
                if (
                    embedded_workbook_blobs(block)
                    and next_text == "点击图片可查看完整电子表格"
                ):
                    index += 2
                    continue
            if entry:
                toc.append(entry)

        elif isinstance(block, Table):
            html_blocks.append(render_table(block))

        index += 1

    title = toc[0]["text"] if toc else doc_info["label"]
    return {
        "id": doc_id,
        "label": doc_info["label"],
        "title": title,
        "body": "\n".join(html_blocks),
        "toc": toc,
    }


def nav_html(report: dict[str, object]) -> str:
    links = []
    for item in report["toc"]:
        level = int(item["level"])
        label = esc(item["text"])
        links.append(
            f'<a class="toc-link level-{level}" href="#{esc(item["id"])}" '
            f'data-report-link="{esc(report["id"])}">{label}</a>'
        )
    return (
        f'<nav class="toc-panel{" active" if report["id"] == "main" else ""}" '
        f'aria-label="{esc(report["label"])}目录" data-toc-panel="{esc(report["id"])}">'
        + "".join(links)
        + "</nav>"
    )


def article_html(report: dict[str, object]) -> str:
    active = " active" if report["id"] == "main" else ""
    return (
        f'<article class="report{active}" data-report="{esc(report["id"])}" '
        f'aria-labelledby="{esc(report["id"])}-title">'
        f'{report["body"]}'
        "</article>"
    )


def build_page(reports: list[dict[str, object]]) -> str:
    tabs = "".join(
        f'<button class="report-tab{" active" if report["id"] == "main" else ""}" '
        f'type="button" data-report-tab="{esc(report["id"])}">{esc(report["label"])}</button>'
        for report in reports
    )
    navs = "\n".join(nav_html(report) for report in reports)
    articles = "\n".join(article_html(report) for report in reports)

    return f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>普林格新四核轮动模型投资报告</title>
  <style>
    :root {{
      --bg: #f6f7f9;
      --paper: #ffffff;
      --ink: #202124;
      --muted: #667085;
      --line: #d9dee7;
      --accent: #a43f32;
      --accent-2: #176d69;
      --accent-soft: #fff1ed;
      --shadow: 0 22px 55px rgba(21, 31, 48, 0.10);
    }}

    * {{
      box-sizing: border-box;
    }}

    html {{
      scroll-behavior: smooth;
    }}

    body {{
      margin: 0;
      color: var(--ink);
      background:
        linear-gradient(135deg, rgba(164, 63, 50, 0.07), transparent 34%),
        linear-gradient(315deg, rgba(23, 109, 105, 0.08), transparent 32%),
        var(--bg);
      font-family: "Microsoft YaHei", "PingFang SC", "Noto Sans CJK SC", Arial, sans-serif;
      letter-spacing: 0;
    }}

    a {{
      color: inherit;
    }}

    .app-shell {{
      display: grid;
      grid-template-columns: 320px minmax(0, 1fr);
      min-height: 100vh;
    }}

    .sidebar {{
      position: sticky;
      top: 0;
      height: 100vh;
      padding: 28px 22px;
      overflow: auto;
      background: rgba(255, 255, 255, 0.86);
      border-right: 1px solid var(--line);
      backdrop-filter: blur(18px);
    }}

    .brand-title {{
      margin: 0 0 8px;
      font-size: 22px;
      line-height: 1.28;
      font-weight: 800;
    }}

    .brand-subtitle {{
      margin: 0 0 20px;
      color: var(--muted);
      font-size: 13px;
      line-height: 1.6;
    }}

    .report-switcher {{
      display: grid;
      gap: 8px;
      margin-bottom: 24px;
    }}

    .report-tab {{
      width: 100%;
      min-height: 42px;
      padding: 0 12px;
      border: 1px solid var(--line);
      border-radius: 8px;
      background: #fff;
      color: #344054;
      font: inherit;
      font-size: 14px;
      font-weight: 700;
      text-align: left;
      cursor: pointer;
    }}

    .report-tab.active {{
      border-color: rgba(164, 63, 50, 0.38);
      background: var(--accent-soft);
      color: var(--accent);
    }}

    .toc-title {{
      margin: 0 0 10px;
      color: #475467;
      font-size: 13px;
      font-weight: 800;
    }}

    .toc-panel {{
      display: none;
      padding-bottom: 28px;
    }}

    .toc-panel.active {{
      display: grid;
      gap: 2px;
    }}

    .toc-link {{
      display: block;
      padding: 7px 8px;
      border-radius: 6px;
      color: #4b5565;
      font-size: 13px;
      line-height: 1.45;
      text-decoration: none;
    }}

    .toc-link:hover,
    .toc-link.active {{
      background: #eef7f6;
      color: var(--accent-2);
    }}

    .toc-link.level-0 {{
      margin-top: 4px;
      color: #111827;
      font-weight: 800;
    }}

    .toc-link.level-2 {{
      padding-left: 18px;
    }}

    .toc-link.level-3 {{
      padding-left: 30px;
      font-size: 12px;
    }}

    .toc-link.level-4 {{
      padding-left: 42px;
      font-size: 12px;
      color: #667085;
    }}

    .toc-link.level-5 {{
      padding-left: 54px;
      font-size: 12px;
      color: #7a8494;
    }}

    .content {{
      min-width: 0;
      padding: 38px clamp(18px, 4vw, 58px) 72px;
    }}

    .report {{
      display: none;
      width: min(100%, 980px);
      margin: 0 auto;
      padding: clamp(28px, 5vw, 62px);
      background: var(--paper);
      border: 1px solid rgba(217, 222, 231, 0.9);
      box-shadow: var(--shadow);
    }}

    .report.active {{
      display: block;
    }}

    .report-title {{
      margin: 0 0 32px;
      padding-bottom: 18px;
      border-bottom: 2px solid #1f2937;
      font-size: clamp(30px, 4vw, 46px);
      line-height: 1.18;
      font-weight: 900;
      color: #111827;
    }}

    h2,
    h3,
    h4,
    h5 {{
      scroll-margin-top: 34px;
      color: #101828;
      letter-spacing: 0;
    }}

    h2 {{
      margin: 44px 0 18px;
      padding-left: 14px;
      border-left: 5px solid var(--accent);
      font-size: 26px;
      line-height: 1.28;
      font-weight: 850;
    }}

    h3 {{
      margin: 32px 0 14px;
      font-size: 22px;
      line-height: 1.35;
      font-weight: 800;
    }}

    h4 {{
      margin: 24px 0 10px;
      font-size: 19px;
      line-height: 1.4;
      font-weight: 800;
      color: #27364a;
    }}

    h5 {{
      margin: 20px 0 8px;
      font-size: 17px;
      line-height: 1.45;
      font-weight: 800;
      color: #344054;
    }}

    .report-paragraph {{
      margin: 10px 0;
      font-size: 16px;
      line-height: 1.9;
      text-align: justify;
    }}

    .align-center {{
      text-align: center;
    }}

    .align-right {{
      text-align: right;
    }}

    .align-justify {{
      text-align: justify;
    }}

    .list-paragraph {{
      position: relative;
      padding-left: calc(22px + var(--list-indent, 0) * 18px);
      text-align: left;
    }}

    .list-paragraph::before {{
      content: "";
      position: absolute;
      left: calc(var(--list-indent, 0) * 18px);
      top: 0.88em;
      width: 6px;
      height: 6px;
      border-radius: 50%;
      background: var(--accent-2);
    }}

    .jump-callout {{
      margin: 26px 0 30px;
      padding: 18px;
      border: 1px solid rgba(164, 63, 50, 0.22);
      border-left: 5px solid var(--accent);
      background: var(--accent-soft);
    }}

    .jump-copy {{
      margin-bottom: 13px;
      font-size: 16px;
      line-height: 1.7;
      font-weight: 800;
      color: #7a2e24;
    }}

    .jump-button {{
      display: inline-flex;
      align-items: center;
      gap: 10px;
      min-height: 40px;
      padding: 0 14px;
      border-radius: 8px;
      background: var(--accent);
      color: #fff;
      font-size: 14px;
      font-weight: 800;
      text-decoration: none;
    }}

    .jump-button::after {{
      content: "";
      width: 8px;
      height: 8px;
      border-top: 2px solid currentColor;
      border-right: 2px solid currentColor;
      transform: rotate(45deg);
    }}

    .report-figure {{
      margin: 24px 0;
      text-align: center;
    }}

    .image-zoom {{
      display: inline-flex;
      width: min(100%, var(--image-width, 900px));
      padding: 0;
      border: 0;
      border-radius: 8px;
      background: transparent;
      cursor: zoom-in;
    }}

    .image-zoom img {{
      display: block;
      width: 100%;
      height: auto;
      border: 1px solid var(--line);
      border-radius: 8px;
      background: #fff;
      box-shadow: 0 12px 30px rgba(16, 24, 40, 0.10);
    }}

    .table-wrap {{
      margin: 22px 0;
      overflow-x: auto;
      border: 1px solid var(--line);
      border-radius: 8px;
    }}

    table {{
      width: 100%;
      border-collapse: collapse;
      min-width: 680px;
      font-size: 15px;
      line-height: 1.65;
    }}

    th,
    td {{
      padding: 12px 14px;
      border-bottom: 1px solid var(--line);
      border-right: 1px solid var(--line);
      vertical-align: top;
      text-align: left;
    }}

    th:last-child,
    td:last-child {{
      border-right: 0;
    }}

    tr:last-child td {{
      border-bottom: 0;
    }}

    th {{
      background: #f1f5f9;
      font-weight: 850;
    }}

    .embedded-workbook {{
      margin: 24px 0 30px;
      border: 1px solid #cfd8e3;
      border-radius: 8px;
      background: #ffffff;
      overflow: hidden;
    }}

    .embedded-workbook-head {{
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 14px;
      min-height: 44px;
      padding: 0 14px;
      border-bottom: 1px solid #cfd8e3;
      background: #eef7f6;
      color: #174f4d;
      font-size: 14px;
      font-weight: 850;
    }}

    .embedded-workbook-head a {{
      color: var(--accent);
      font-size: 13px;
      text-decoration: none;
      white-space: nowrap;
    }}

    .embedded-sheet {{
      padding: 14px;
    }}

    .embedded-sheet + .embedded-sheet {{
      border-top: 1px solid #e4e7ec;
    }}

    .embedded-sheet-name {{
      margin-bottom: 8px;
      color: #667085;
      font-size: 12px;
      font-weight: 800;
    }}

    .embedded-table-wrap {{
      margin: 0;
      max-height: 62vh;
      border-radius: 6px;
      background: #fff;
    }}

    .embedded-table {{
      width: auto;
      min-width: 100%;
      font-size: 14px;
      line-height: 1.55;
    }}

    .embedded-table th {{
      position: sticky;
      top: 0;
      z-index: 1;
      background: #edf2f7;
      color: #1f2937;
      white-space: nowrap;
    }}

    .embedded-table td,
    .embedded-table th {{
      min-width: 72px;
      max-width: 360px;
      word-break: break-word;
    }}

    .lightbox {{
      position: fixed;
      inset: 0;
      z-index: 20;
      display: none;
      place-items: center;
      padding: 36px;
      background: rgba(10, 16, 26, 0.86);
    }}

    .lightbox.active {{
      display: grid;
    }}

    .lightbox img {{
      max-width: 96vw;
      max-height: 90vh;
      object-fit: contain;
      background: #fff;
      border-radius: 8px;
    }}

    .lightbox-close {{
      position: fixed;
      top: 18px;
      right: 18px;
      width: 42px;
      height: 42px;
      border: 1px solid rgba(255, 255, 255, 0.35);
      border-radius: 8px;
      background: rgba(255, 255, 255, 0.14);
      color: #fff;
      font-size: 24px;
      line-height: 1;
      cursor: pointer;
    }}

    @media (max-width: 980px) {{
      .app-shell {{
        display: block;
      }}

      .sidebar {{
        position: sticky;
        top: 0;
        z-index: 10;
        height: auto;
        max-height: 55vh;
        padding: 18px;
      }}

      .brand-title {{
        font-size: 20px;
      }}

      .report-switcher {{
        grid-template-columns: repeat(auto-fit, minmax(112px, 1fr));
      }}

      .report-tab {{
        text-align: center;
      }}

      .content {{
        padding: 20px 12px 52px;
      }}

      .report {{
        padding: 26px 18px 36px;
      }}
    }}

    @media (max-width: 640px) {{
      .report-switcher {{
        grid-template-columns: 1fr;
      }}

      .report-title {{
        font-size: 29px;
      }}

      h2 {{
        font-size: 23px;
      }}

      h3 {{
        font-size: 20px;
      }}

      .report-paragraph,
      table {{
        font-size: 15px;
      }}
    }}
  </style>
</head>
<body>
  <div class="app-shell">
    <aside class="sidebar">
      <div class="brand-title">普林格轮动报告</div>
      <p class="brand-subtitle">四份报告整合呈现</p>
      <div class="report-switcher" aria-label="报告切换">
        {tabs}
      </div>
      <div class="toc-wrap">
        <div class="toc-title">目录</div>
        {navs}
      </div>
    </aside>
    <main class="content">
      {articles}
    </main>
  </div>
  <div class="lightbox" id="lightbox" aria-hidden="true">
    <button class="lightbox-close" type="button" aria-label="关闭">×</button>
    <img alt="">
  </div>
  <script>
    const tabs = [...document.querySelectorAll("[data-report-tab]")];
    const reports = [...document.querySelectorAll("[data-report]")];
    const panels = [...document.querySelectorAll("[data-toc-panel]")];
    const tocLinks = [...document.querySelectorAll("[data-report-link]")];

    function setActiveReport(id, updateHash = true) {{
      tabs.forEach((tab) => tab.classList.toggle("active", tab.dataset.reportTab === id));
      reports.forEach((report) => report.classList.toggle("active", report.dataset.report === id));
      panels.forEach((panel) => panel.classList.toggle("active", panel.dataset.tocPanel === id));
      if (updateHash) {{
        history.replaceState(null, "", "#" + id + "-title");
      }}
    }}

    function activateHash(hash) {{
      const targetId = hash ? hash.replace(/^#/, "") : "main-title";
      const target = document.getElementById(targetId);
      if (!target) {{
        setActiveReport("main", false);
        return;
      }}
      const report = target.closest("[data-report]");
      if (report) {{
        setActiveReport(report.dataset.report, false);
      }}
      requestAnimationFrame(() => target.scrollIntoView({{ block: "start" }}));
    }}

    tabs.forEach((tab) => {{
      tab.addEventListener("click", () => {{
        const id = tab.dataset.reportTab;
        setActiveReport(id);
        document.getElementById(id + "-title")?.scrollIntoView({{ block: "start" }});
      }});
    }});

    tocLinks.forEach((link) => {{
      link.addEventListener("click", (event) => {{
        const id = link.dataset.reportLink;
        setActiveReport(id, false);
        const target = document.querySelector(link.getAttribute("href"));
        if (target) {{
          event.preventDefault();
          history.replaceState(null, "", link.getAttribute("href"));
          target.scrollIntoView({{ block: "start" }});
        }}
      }});
    }});

    document.querySelectorAll(".cross-report-link").forEach((link) => {{
      link.addEventListener("click", (event) => {{
        event.preventDefault();
        const id = link.dataset.targetReport;
        setActiveReport(id, false);
        const target = document.querySelector(link.getAttribute("href"));
        if (target) {{
          history.replaceState(null, "", link.getAttribute("href"));
          target.scrollIntoView({{ block: "start" }});
        }}
      }});
    }});

    const lightbox = document.getElementById("lightbox");
    const lightboxImage = lightbox.querySelector("img");
    document.querySelectorAll(".image-zoom").forEach((button) => {{
      button.addEventListener("click", () => {{
        lightboxImage.src = button.dataset.imageSrc;
        lightbox.classList.add("active");
        lightbox.setAttribute("aria-hidden", "false");
      }});
    }});
    function closeLightbox() {{
      lightbox.classList.remove("active");
      lightbox.setAttribute("aria-hidden", "true");
      lightboxImage.removeAttribute("src");
    }}
    lightbox.querySelector(".lightbox-close").addEventListener("click", closeLightbox);
    lightbox.addEventListener("click", (event) => {{
      if (event.target === lightbox) closeLightbox();
    }});
    window.addEventListener("keydown", (event) => {{
      if (event.key === "Escape" && lightbox.classList.contains("active")) closeLightbox();
    }});

    window.addEventListener("hashchange", () => activateHash(location.hash));
    activateHash(location.hash || "#main-title");
  </script>
</body>
</html>
"""


def main() -> None:
    ASSET_DIR.mkdir(parents=True, exist_ok=True)
    reports = [convert_doc(item) for item in DOCS]
    OUTPUT_FILE.write_text(build_page(reports), encoding="utf-8")
    print(f"Generated {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
